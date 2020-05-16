import openpyxl

wb = openpyxl.load_workbook('automate_online-materials/example.xlsx')
sheet = wb.active
print(sheet['A1'])
print(sheet['A1'].value)
c = sheet['B1']
print(c.value)

print('Row {}, Column {} is {}.'.format(str(c.row), c.column, c.value))
print('Cell {} is {}.'.format(c.coordinate, c.value))

print(sheet.cell(row=1, column=2))
print(sheet.cell(row=1, column=2).value)

for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)
